from PySide6.QtWidgets import (
    QTableWidgetItem, QMessageBox, QFileDialog
)
from app.models.report_model import ReportModel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

class ReportController:
    def __init__(self, view):
        self.view = view
        self.selected_id = None

        self.load_data()

        self.view.table.cellClicked.connect(self.select_row)
        self.view.btn_detail.clicked.connect(self.show_detail)
        self.view.btn_export.clicked.connect(self.export_to_excel)

    def load_data(self):
        data = ReportModel.get_transactions()
        self.view.table.setRowCount(len(data))

        total_penjualan = 0

        for row, item in enumerate(data):
            self.view.table.setItem(row, 0, QTableWidgetItem(str(item["id"])))
            self.view.table.setItem(row, 1, QTableWidgetItem(item["date"]))
            self.view.table.setItem(row, 2, QTableWidgetItem(str(item["total"])))

            total_penjualan += item["total"]

        self.view.label_total.setText(
            f"Total Penjualan: {total_penjualan}"
        )

    def select_row(self, row, col):
        self.selected_id = int(
            self.view.table.item(row, 0).text()
        )

    def show_detail(self):
        if not self.selected_id:
            QMessageBox.warning(
                self.view,
                "Error",
                "Pilih transaksi terlebih dahulu"
            )
            return

        details = ReportModel.get_transaction_detail(self.selected_id)

        detail_text = ""
        for d in details:
            detail_text += (
                f"{d['name']} | "
                f"Qty: {d['quantity']} | "
                f"Subtotal: {d['subtotal']}\n"
            )

        QMessageBox.information(
            self.view,
            "Detail Transaksi",
            detail_text
        )

    def export_to_excel(self):
        """Export laporan penjualan ke file Excel"""
        try:
            # Dialog untuk memilih lokasi save file
            file_path, _ = QFileDialog.getSaveFileName(
                self.view,
                "Export Laporan Penjualan",
                f"Laporan_Penjualan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx)"
            )

            if not file_path:
                return  # User cancelled

            # Ambil data transaksi
            data = ReportModel.get_transactions()

            # Buat workbook baru
            wb = Workbook()
            ws = wb.active
            ws.title = "Laporan Penjualan"

            # Style untuk header
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Tulis judul
            ws.merge_cells('A1:D1')
            ws['A1'] = "LAPORAN PENJUALAN"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = header_alignment

            # Tulis tanggal export
            ws.merge_cells('A2:D2')
            ws['A2'] = f"Tanggal Export: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
            ws['A2'].alignment = Alignment(horizontal="center")

            # Header kolom
            headers = ["No", "ID Transaksi", "Tanggal", "Total"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # Tulis data
            total_penjualan = 0
            for idx, item in enumerate(data, start=1):
                ws.cell(row=idx+4, column=1, value=idx)
                ws.cell(row=idx+4, column=2, value=item["id"])
                ws.cell(row=idx+4, column=3, value=item["date"])
                ws.cell(row=idx+4, column=4, value=item["total"])
                total_penjualan += item["total"]

            # Baris total
            total_row = len(data) + 5
            ws.merge_cells(f'A{total_row}:C{total_row}')
            total_cell = ws.cell(row=total_row, column=1)
            total_cell.value = "TOTAL PENJUALAN"
            total_cell.font = Font(bold=True, size=12)
            total_cell.alignment = Alignment(horizontal="right")
            
            total_value_cell = ws.cell(row=total_row, column=4)
            total_value_cell.value = total_penjualan
            total_value_cell.font = Font(bold=True, size=12)
            total_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            total_value_cell.fill = total_fill

            # Atur lebar kolom
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 15

            # Simpan file
            wb.save(file_path)

            QMessageBox.information(
                self.view,
                "Sukses",
                f"Laporan berhasil di-export ke:\n{file_path}"
            )

        except Exception as e:
            QMessageBox.critical(
                self.view,
                "Error",
                f"Gagal export ke Excel: {str(e)}"
            )
